import datetime as dt
import os
from flask import flash, redirect, request, url_for, render_template
from sqlalchemy.orm import aliased
from werkzeug.utils import secure_filename
from library.python_files.Database_tables import Book, Book_Copies, User, Book_Status
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


def create_excel_file(all_data):
    book_records = {}
    workbook = Workbook()

    # Get the active worksheet (the first sheet by default)
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = ['Name', 'Book_Title', 'Book_Reserve_Date', 'Book_Approve_Date', 'Book_Return_Date',
               'Book_Approval_Status']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add book data to the worksheet
    for row_num, data in enumerate(all_data, 2):
        for col_num, header in enumerate(headers, 1):
            # approve_date = dt.datetime.strptime(book.current_date, "%m/%d/%Y, %H:%M:%S")
            # return_date = approve_date + dt.timedelta(days=7)

            book_records = {
                'Name': data.first_name + data.last_name,
                'Book_Title': data.title,
                'Book_Reserve_Date': data.reserve_book_date,
                'Book_Approve_Date': data.approve_book_date,
                'Book_Return_Date': data.return_book_date,
                'Book_Approval_Status': data.book_approvals_status,

            }
            worksheet.cell(row_num, col_num, book_records[header])

        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            column_width = max(len(header), max(len(str(book)) for book in book_records))
            worksheet.column_dimensions[column_letter].width = column_width + 2
            # # Save the workbook to a file
        workbook.save('./static/reports/book_database.xlsx')

    return "Excel file book_database created successfully!"


def create_pdf_file(all_data):
    doc = SimpleDocTemplate("./static/reports/book_database.pdf", pagesize=letter)

    # Create a list to hold the table data
    table_data = []

    # Define table column headers
    table_data.append(["Name", "Book Title", "Reserve Date", "Approve Date", "Return Date", "Approval Status"])

    # Populate the table with data from all_data
    for data in all_data:
        row = [
            f"{data.first_name} {data.last_name}",
            data.title,
            data.reserve_book_date,
            data.approve_book_date,
            data.return_book_date,
            data.book_approvals_status
        ]
        table_data.append(row)

    # Create the table and define table style
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    # Build the PDF document
    doc.build([table])
    return "PDF file created successfully!"


class Reports:
    def __init__(self, session, app):
        self.session = session
        self.app = app

    def generate_reports(self, file_type):
        bs = aliased(Book_Status, name='bs')
        users_with_books_data = self.session.query(User.first_name, User.last_name, Book.title,
                                                   bs.book_approvals_status, bs.reserve_book_date,
                                                   bs.return_book_date, bs.approve_book_date).join(
            Book).join(User).filter(bs.book_id == Book.id).all()

        #     books =session.query(Book,Book_Status)
        if file_type == 'excel':
            is_created = create_excel_file(users_with_books_data)
            return 'excel'
        elif file_type == 'pdf':
            is_created = create_pdf_file(users_with_books_data)
            return 'pdf'
        else:
            return "Invalid file type. Use 'excel' or 'pdf'."

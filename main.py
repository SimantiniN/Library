import datetime as dt
from functools import wraps
from flask import Flask, render_template, redirect, url_for, flash, request, send_from_directory
from flask_bootstrap import Bootstrap
from flask_ckeditor import CKEditor
from werkzeug.security import generate_password_hash, check_password_hash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine, ForeignKey, text
from sqlalchemy.orm import relationship, Session, aliased
from flask_login import UserMixin, login_user, LoginManager, login_required, current_user, logout_user
from forms import RegisterForm, LoginForm, BookDetailsForm, ReviewForm
import os
from werkzeug.utils import secure_filename
from flask_bcrypt import Bcrypt
from flask_gravatar import Gravatar
from flask import abort
from functools import wraps
import pandas as pd
from reportlab.pdfgen import canvas
from openpyxl import Workbook

import xlsxwriter
from openpyxl.utils import get_column_letter
import xlwt
from itsdangerous import URLSafeTimedSerializer
from flask_mail import Mail, Message

app = Flask(__name__)
app.config['SECRET_KEY'] = "h6a6r1h3a27r8386ma99ha1d3va"
app.config['SECURITY_PASSWORD_SALT'] = "Omnamhashivay"
bootstrap = Bootstrap(app)
ckeditor = CKEditor(app)
bcrypt = Bcrypt(app)

# Configure Flask-Mail settings
app.config['MAIL_SERVER'] = 'your_mail_server'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'your_mail_username'
app.config['MAIL_PASSWORD'] = 'your_mail_password'

# CONNECT TO DB
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///Library.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
session = Session()

# upload the image /file
UPLOAD_FOLDER = './static/books'  # where to store uploaded files
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# gracator
gravatar = Gravatar(app, size=75, rating='g', default='robohash', force_default=False, force_lower=False, use_ssl=False,
                    base_url=None)


# CONFIGURE TABLES
class User(UserMixin, db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(100))
    last_name = db.Column(db.String(100))
    email = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(100))
    all_users_status = relationship("Book_Status", back_populates="user_status")
    users_review = relationship("Book_Review", back_populates="user_review")
    all_users_role = relationship("User_Role", back_populates="user_role")


class User_Role(db.Model):
    __tablename__ = 'users_role'
    id = db.Column(db.Integer, primary_key=True)
    role = db.Column(db.String(50))
    description = db.Column(db.String(150))
    user_id = db.Column(db.Integer, db.ForeignKey(User.id))
    user_role = relationship("User", back_populates="all_users_role")


class Book(db.Model):
    __tablename__ = 'books'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(150))
    sub_title = db.Column(db.String(150))
    author = db.Column(db.String(150))
    brand_name = db.Column(db.String(100))
    number_of_pages = db.Column(db.Integer)
    book_contains = db.Column(db.String(100))
    isbn = db.Column(db.String(100), unique=True)
    image = db.Column(db.String(100), unique=False)
    no_of_copies = db.Column(db.Integer)
    all_book_status = relationship("Book_Status", back_populates="book_status_all")
    book_reviews = relationship("Book_Review", back_populates='book_review')
    book_copy = relationship("Book_Copies", back_populates="book_copies")


class Book_Copies(db.Model):
    __tablename__ = 'book_copies'
    id = db.Column(db.Integer, primary_key=True)
    book_id = db.Column(db.Integer, db.ForeignKey(Book.id))
    book_copies = relationship("Book", back_populates="book_copy")
    remaining_copies = db.Column(db.Integer, unique=False)
    # book_avialabity


class Book_Status(db.Model):
    __tablename__ = 'books_status'
    id = db.Column(db.Integer, primary_key=True)
    book_approvals_status = db.Column(db.String(50))
    reserve_book_date = db.Column(db.String)
    approve_book_date = db.Column(db.String)
    return_book_date = db.Column(db.String)
    book_id = db.Column(db.Integer, db.ForeignKey(Book.id))
    book_status_all = relationship("Book", back_populates="all_book_status")
    user_id = db.Column(db.Integer, db.ForeignKey(User.id))
    user_status = relationship("User", back_populates="all_users_status")


class Book_Review(db.Model):
    __tablename__ = "book_review"
    id = db.Column(db.Integer, primary_key=True)
    review_text = db.Column(db.Text, nullable=False)
    book_id = db.Column(db.Integer, db.ForeignKey(Book.id))
    book_review = relationship("Book", back_populates="book_reviews")
    user_id = db.Column(db.Integer, db.ForeignKey(User.id))
    user_review = relationship("User", back_populates="users_review")


with app.app_context():
    db.create_all()

# login Authentication
login_manager = LoginManager()
login_manager.init_app(app)


@login_manager.user_loader
def load_user(user_id):
    return db.session.query(User).get(user_id)  # User.query.get(user_id)


# admin permission
def admin_only(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.id != 1 or not current_user.is_authenticated:
            return abort(403)
        return f(*args, **kwargs)

    return decorated_function


# functions to upload image
@app.route('/uploads/<filename>')
def get_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'],
                               filename)


def allowed_file(filename):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# @app.route("/reserve_book_count/<int:user_id>")
# @login_required
def reserve_books_count(user_id):
    # if current_user.is_authenticated:
    # total_reserve_books =db.session.query(Book_Status)
    total_reserve_books = db.session.query(Book_Status).filter(Book_Status.user_id == user_id).count()
    if not total_reserve_books:
        flash("Currently no books reserve from you!Hurry UP!!\n")
        return 0

    else:
        return total_reserve_books


# return redirect(url_for('home'))
@app.route('/')
def home():
    total_reserve_books = 0
    current_date = dt.datetime.today().strftime("%B %d,%Y")
    book_data = db.session.query(Book).all()

    # total_book_copies = db.session.query(Book_Copies).all()

    if not book_data:
        flash("Contact Admin no books are currently available.\n")
    if current_user.is_authenticated:
        total_reserve_books = reserve_books_count(current_user.id)
    #     total_reserve_books = Book.query.filter(Book.user_id == current_user.id).count()
    #     if not total_reserve_books:
    #         flash("Currently no books reserve from you!Hurry UP!!\n")
    # else:
    #     total_reserve_books = 0
    return render_template("index.html", date=current_date, books=book_data, reserve_books_count=total_reserve_books,
                           logged_in=current_user.is_authenticated)  # , total_book_copies=total_book_copies)


@app.route('/register_user', methods=['GET', 'POST'])
def register_user():
    registration_form = RegisterForm()
    if registration_form.validate_on_submit():
        if User.query.filter_by(email=request.form.get('email')).first():
            # User already exists
            flash("You've already signed up with that email, log in instead!\n")
            return redirect(url_for('login'))
        hash_and_salted_password = bcrypt.generate_password_hash(request.form.get('password'), 10)
        # #generate_password_hash(
        #     request.form.get('password'),
        #     method='pbkdf2:sha256',
        #     salt_length=8
        # )
        new_user = User()
        new_user.first_name = registration_form.firstname.data
        new_user.last_name = registration_form.lastname.data
        new_user.email = registration_form.email.data
        new_user.password = hash_and_salted_password
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for("login"))
    return render_template("registration.html", form=registration_form)


@app.route('/login', methods=['GET', 'POST'])
def login():
    login_form = LoginForm()
    if login_form.validate_on_submit():
        email = login_form.email.data
        password = login_form.password.data
        user = User.query.filter_by(email=email).first()
        if user and bcrypt.check_password_hash(user.password, password):
            login_user(user)
            return redirect(url_for('home'))
        else:
            flash("Username or Password incorrect!Please try again.\n")
    return render_template('login.html', form=login_form, logged_in=current_user.is_authenticated)


@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('home'))


@app.route('/add_book', methods=['GET', 'POST'])
@admin_only
@login_required
def add_book():
    return render_template('add_book.html')


@login_required
@admin_only
@app.route('/add_book_details', methods=['GET', 'POST'])
def add_book_details():
    book_exist = False
    books = Book.query.all()
    book_form = BookDetailsForm()
    if book_form.validate_on_submit():
        for book in books:
            if book.title == book_form.title.data.title():
                book_exist = True
                if book_exist:
                    flash('Book is already exist')
                    book_exist = False
                    return redirect(url_for('home'))

        file = book_form.image.data
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            file_url = url_for('get_file', filename=filename)
            book = Book()
            book.title = book_form.title.data.title()
            book.sub_title = book_form.subtitle.data
            book.author = book_form.author.data
            book.brand_name = book_form.brand_name.data
            book.number_of_pages = book_form.number_of_pages.data
            book.book_contains = book_form.book_contains.data
            book.isbn = book_form.isbn.data
            book.image = file_url
            book.no_of_copies = book_form.no_of_copies.data
            db.session.add(book)
            db.session.commit()
            bs = Book_Status()
            total_book_copies = Book_Copies()
            total_book_copies.book_id = book.id
            if book.no_of_copies > 0:
                bs = Book_Status()
                bs.book_id = book.id
                # bs.user_id = current_user.id
                bs.book_approvals_status = "Available"
                total_book_copies.remaining_copies = book.no_of_copies
            else:
                total_book_copies.remaining_copies = 0
                total_book_copies.book_id = book.id
                bs.book_approvals_status = "Waiting"
            db.session.add(bs)
            db.session.add(total_book_copies)
            db.session.commit()
            return redirect(url_for('home'))
    return render_template('add_book_details.html', form=book_form, logged_in=current_user.is_authenticated)


# @app.route("/book-status/<int:book_id>")
# @login_required
# def book_status(book_id):
@app.route("/book_review/<int:book_id>", methods=['GET', 'POST'])
@login_required
def add_book_review(book_id):
    review_form = ReviewForm()
    try:
        book = db.session.query(Book.id, Book.image, Book_Status.book_approvals_status).join(Book_Status).filter(
            Book_Status.book_id == book_id).first()
        if book is None:
            # Handle the case where the book with the given ID doesn't exist
            return "Book not found", 404
        if review_form.validate_on_submit():
            if not current_user.is_authenticated:
                flash("You need to login or register to comment.\n")
                return redirect(url_for("login"))
            new_review = Review(
                review_text=review_form.review_text.data,
                # book_id = book_id,
                # user_id =current_user.id,
                user_review=current_user,
                book_review=requested_book
            )
            db.session.add(new_review)
            db.session.commit()

    except Exception as e:
        db.session.rollback()

    return render_template("book_reviews.html", book=book, form=review_form,
                           current_user=current_user,
                           logged_in=current_user.is_authenticated)


@app.route("/edit_book/<int:book_id>", methods=['GET', 'POST'])
@login_required
@admin_only
def edit_book(book_id):
    book_to_edit = db.session.get_or_404(Book, book_id)
    edit_form = BookDetailsForm(

        title=book_to_edit.title,
        author=book_to_edit.author
    )
    if edit_form.validate_on_submit():
        book_to_edit.title = edit_form.title.data
        book_to_edit.author = edit_form.author.data
        book_to_edit.no_of_copies = edit_form.no_of_copies.data
        file = edit_form.image.data
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            file_url = url_for('get_file', filename=filename)
            book_to_edit.image = file_url
            book_to_edit.date = dt.datetime.now()
            # db.session.add(book_to_edit)
            db.session.commit()
        return redirect(url_for('home'))
    return render_template('add_book_details.html', form=edit_form, is_edit=True,
                           current_user=current_user.is_authenticated)


@app.route("/delete/<int:book_id>")
@login_required
@admin_only
def delete_book(book_id):
    book_to_delete = Book.query.get(book_id)
    bs = Book_Status()
    bs.book_approvals_status = "Not available"
    db.session.delete(book_to_delete)
    db.session.commit()
    return redirect(url_for('home'))


@app.route("/reserve_book/<int:book_id>")
@login_required
def reserve_book(book_id):
    book, book_status, book_copies = db.session.query(Book, Book_Status, Book_Copies).join(Book_Status).join(
        Book_Copies).filter(Book.id == book_id).first()

    if book_copies.remaining_copies > 0:
        if not book:
            flash('Book not found!\n', 'error')
            return redirect((url_for('home')))
        elif book_status.book_approvals_status == 'Reserved' and book_status.user_id == current_user.id:
            flash('Book is already reserved!\n', 'error\n')
            return redirect((url_for('home')))
        elif book_status.book_approvals_status == 'Approved' and book_status.user_id == current_user.id:
            flash('Book is already Approved!\n', 'error\n')
        else:
            if book_status.book_approvals_status == "Available":
                book_status.book_approvals_status = "Reserved"
                book_copies.remaining_copies = book_copies.remaining_copies - 1
                book_status.reserve_book_date = dt.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
                book_status.user_id = current_user.id
            else:
                if book_status.book_approvals_status == "Reserved" and book_status.user_id != current_user.id:
                    book_status.book_approvals_status = "Reserved"
                    different_user_book_status = Book_Status(book_id=book.id, user_id=current_user.id,
                                                             reserve_book_date=dt.datetime.now().strftime(
                                                                 "%m/%d/%Y, %H:%M:%S"),
                                                             book_approvals_status="Reserved")
                    db.session.add(different_user_book_status)
                    book_copies.remaining_copies = book_copies.remaining_copies - 1
                    book_status.reserve_book_date = dt.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
                    book_status.user_id = current_user.id

            db.session.commit()
            flash(f'Book "{book.title}" reserved successfully!\n', 'success\n')
    else:
        if not book_status.book_approvals_status == "Approved":
            different_user_book_status = Book_Status(book_id=book.id, user_id=current_user.id,
                                                     reserve_book_date=dt.datetime.now().strftime("%m/%d/%Y, %H:%M:%S"),
                                                     book_approvals_status="Waiting")
            db.session.add(different_user_book_status)
            book_status.user_id = current_user.id
            db.session.commit()
            flash('Currently not available !!You are in In waiting list')
    # else:
    #     if book_status.book_approvals_status== "Available":
    #         book_status.book_approvals_status = "Reserved"
    #         total_book_copies.remaining_copies = total_book_copies.remaining_copies - 1
    #         # book_copies_count = Book_Copies()
    #         # book_copies_count.remaining_copies = total_book_count
    #         book.user_id = current_user.id
    #         db.session.commit()
    #         flash(f'Book "{book_details.title}" reserved successfully!\n', 'success\n')
    # # except Exception as e:
    # #     db.session.rollback()
    return redirect((url_for('home')))


@app.route('/dereserve_book/<int:book_id>/<int:user_id>')
@login_required
def dereserve_book(book_id, user_id):
    book = db.session.query(Book.title, Book_Status.book_approvals_status, Book_Status.user_id, Book_Status.book_id,
                            Book_Copies.remaining_copies). \
        join(Book_Status).join(Book_Copies). \
        filter(Book.id == book_id).first()
    if book.book_approvals_status == 'Reserved' and book.user_id == current_user.id:
        query = text(
            "UPDATE books_status "
            "SET book_approvals_status = 'Available', user_id = :new_user_id ,reserve_book_date=''"
            "WHERE id = :book_id AND user_id = :user_id AND book_approvals_status = 'Reserved'")


        if book.book_approvals_status == 'Reserved' and book.user_id == current_user.id:
            book.book_approvals_status = "Available"
            Book_Status.user_id = ""
        db.session.execute(query, {"book_id": book_id, "user_id": user_id, "new_user_id": " "})
        db.session.commit()
    flash(f'Book "{book.title}" dereserved successfully!\n', 'success\n')
    # else:
    #     flash('Book is approve So cannot dereserve!\n', 'error\n')

    if book is None:
        flash('Book not found!\n', 'error\n')

    return redirect((url_for('home')))


@app.route('/book_approvals', methods=['GET', 'POST'])
@login_required
@admin_only
def book_approvals():
    bs = aliased(Book_Status, name='bs')
    users_with_books = db.session.query(User.first_name, Book.id, Book.title, bs.book_approvals_status,
                                        bs.return_book_date, bs.approve_book_date).join(
        Book).join(User).filter(bs.book_id == Book.id).all()

    total_reserve_books = reserve_books_count(current_user.id)
    if not users_with_books:
        flash('No books are available to reserved!!')
    return render_template('book_approval.html', users_with_books=users_with_books,
                           reserve_books_count=total_reserve_books, logged_in=current_user.is_authenticated)


@app.route('/approved_book/<int:book_id>', methods=['GET', 'POST'])
@login_required
@admin_only
def approved_book(book_id):
    approved_status = db.session.query(Book_Status).filter_by(id=book_id).first()
    if approved_status.book_approvals_status == 'Reserved':
        approved_status.book_approvals_status = "Approved"
        current_date = dt.datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
        approved_status.approve_book_date = current_date
        approve_date = dt.datetime.strptime(current_date, "%m/%d/%Y, %H:%M:%S")
        approved_status.return_book_date = str(approve_date + dt.timedelta(days=7))
        db.session.commit()
        flash(f'Approved!!return your book on {approved_status.return_book_date}')
        return redirect(url_for('book_approvals'))

    else:
        total_book_copies = Book_Copies()
        total_book_copies.remaining_copies = Book.no_of_copies
        flash("flash('No books reserved!!')")
        return redirect(url_for('home'))


@app.route('/return_book/<int:book_id>', methods=['GET', 'POST'])
@login_required
@admin_only
def return_book(book_id):
    return_bs = db.session.query(Book_Status).filter_by(id=book_id).first()
    if return_bs.book_approvals_status == 'Approved':
        return_bs.book_approvals_status = "Available"
        return_bs.return_date = ""
        return_bs.no_of_copies = return_bs.no_of_copies + 1
        db.session.commit()
        flash(f'Returned')
        return redirect(url_for('book_approvals'))
    else:
        flash("flash('No books approved!!')")
        return redirect(url_for('home'))


def create_excel_file(data):
    book_records = {}
    workbook = Workbook()

    # Get the active worksheet (the first sheet by default)
    worksheet = workbook.active

    # Add headers to the worksheet
    headers = ['Title', 'Author', 'current_date', 'return_date']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Add book data to the worksheet
    for row_num, book in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            # approve_date = dt.datetime.strptime(book.current_date, "%m/%d/%Y, %H:%M:%S")
            # return_date = approve_date + dt.timedelta(days=7)

            book_records = {
                'Title': book.title,
                'Author': book.author,
                'current_date': book.current_date,
                'return_date': book.return_date

            }
            worksheet.cell(row_num, col_num, book_records[header])

        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            column_width = max(len(header), max(len(str(book)) for book in book_records))
            worksheet.column_dimensions[column_letter].width = column_width + 2
            # # Save the workbook to a file
        workbook.save('./static/reports/book_database.xlsx')

    return "Excel file book_database created successfully!"


def create_pdf_file(data):
    # Create a PDF using ReportLab
    pdf_filename = 'output.pdf'
    c = canvas.Canvas(pdf_filename)
    y = 800  # Y-position for starting content

    # Iterate through the data and add it to the PDF
    for item in data:
        content = f"Column1: {item.id}, Column2: {item.title}"
        c.drawString(100, y, content)
        y -= 20

    c.save()

    return "PDF file created successfully!"


@app.route('/create_file/<file_type>')
@admin_only
@login_required
def create_file(file_type):
    data = Book.query.all()
    if file_type == 'excel':
        is_created = create_excel_file(data)
        return render_template('reports.html', msg=is_created, logged_in=current_user.is_authenticated)
    elif file_type == 'pdf':
        return create_pdf_file(data)
    else:
        return "Invalid file type. Use 'excel' or 'pdf'."


@app.route('/download')
@admin_only
@login_required
def download():
    return send_from_directory(directory='static', path='reports/book_database.xlsx')


if __name__ == "__main__":
    app.run(debug=True)

# workbook = xlsxwriter.Workbook('book_database.xlsx')
#
# # Add a worksheet to the workbook
# worksheet = workbook.add_worksheet()
#
# # Add headers to the worksheet
# headers = ['Title', 'Author']
# for col_num, header in enumerate(headers):
#     worksheet.write(0, col_num, header)
#
# # Add book data to the worksheet
# for row_num, book in enumerate(data, 1):
#     for col_num, header in enumerate(headers,0):
#         book_rec = {
#             'Title': book.title,
#             'Author':book.author,
#         }
#         worksheet.write(row_num, col_num, book_rec[header])  # Write 'Title' to the first column (column 0)
#     # worksheet.write(row_num, col_num, book.author)  # Write 'Author' to the second column (column 1)
#
# # Save the workbook
# workbook.close()
# <p>Debug: book = {{ book }}</p>
# <p>Debug: book_approvals_status = {{ book.book_approvals_status }}</p>
